# import tkinter as tk
# from PIL import Image, ImageTk, ImageDraw, ImageFont
# from openpyxl import load_workbook
# import requests
# from io import BytesIO

# EXCEL_FILE = "folder_status.xlsx"

# # === Load Data from Excel ===
# def load_excel_data():
#     wb = load_workbook(EXCEL_FILE)
#     ws = wb.active
#     headers = [cell.value for cell in ws[1]]

#     data = {}
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         folder_name = row[0]
#         status = row[1]
#         input_urls = row[2].split(",") if row[2] else []
#         predicted_urls = row[3].split(",") if row[3] else []
#         data[folder_name] = {
#             "status": status,
#             "input_urls": input_urls,
#             "predicted_urls": predicted_urls
#         }
#     return data, wb, ws

# class ImageViewer:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("URL Image Viewer")
#         self.current_index = 0

#         self.data, self.wb, self.ws = load_excel_data()
#         self.folder_names = list(self.data.keys())

#         # === GUI Layout ===
#         self.search_frame = tk.Frame(root)
#         self.search_frame.pack(fill=tk.X, padx=5, pady=5)

#         self.search_entry = tk.Entry(self.search_frame)
#         self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

#         self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
#         self.search_button.pack(side=tk.LEFT, padx=5)

#         self.canvas = tk.Canvas(root, bg='black')
#         self.canvas.pack(fill=tk.BOTH, expand=True)

#         self.button_frame = tk.Frame(root)
#         self.button_frame.pack(fill=tk.X, pady=10)

#         self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
#         self.prev_button.pack(side=tk.LEFT, padx=10)

#         self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
#         self.next_button.pack(side=tk.RIGHT, padx=10)

#         self.status_label = tk.Label(root, text="Enter Status:")
#         self.status_label.pack(pady=5)

#         self.status_entry = tk.Entry(root)
#         self.status_entry.pack(fill=tk.X, padx=5, pady=5)
#         self.status_entry.bind("<Return>", lambda e: self.save_status())

#         self.save_button = tk.Button(root, text="Save Status", command=self.save_status)
#         self.save_button.pack(pady=5)

#         self.display_images()

#     def load_images_from_urls(self, urls):
#         images = []
#         for url in urls:
#             try:
#                 resp = requests.get(url)
#                 resp.raise_for_status()
#                 img = Image.open(BytesIO(resp.content)).resize((200, 200))
#                 images.append(img)
#             except Exception as e:
#                 print(f"Error loading image from {url}: {e}")
#         return images

#     def label_row(self, text, width):
#         img = Image.new("RGB", (width, 30), color="gray")
#         draw = ImageDraw.Draw(img)
#         font = ImageFont.load_default()
#         bbox = draw.textbbox((0, 0), text, font=font)
#         w = bbox[2] - bbox[0]
#         h = bbox[3] - bbox[1]
#         draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
#         return img

#     def display_images(self):
#         self.canvas.delete("all")
#         folder = self.folder_names[self.current_index]
#         entry = self.data[folder]

#         input_imgs = self.load_images_from_urls(entry["input_urls"])
#         predicted_imgs = self.load_images_from_urls(entry["predicted_urls"])

#         rows = []
#         for label, images in zip(["Input Images", "Predicted Images"], [input_imgs, predicted_imgs]):
#             if images:
#                 row_width = sum(img.width for img in images)
#                 row_height = max(img.height for img in images)
#                 row_img = Image.new("RGB", (row_width, row_height))
#                 x_offset = 0
#                 for img in images:
#                     row_img.paste(img, (x_offset, 0))
#                     x_offset += img.width

#                 label_img = self.label_row(label, row_width)
#                 combined = Image.new("RGB", (row_width, row_height + 30), color="black")
#                 combined.paste(label_img, (0, 0))
#                 combined.paste(row_img, (0, 30))
#                 rows.append(combined)

#         if rows:
#             total_width = max(r.width for r in rows)
#             total_height = sum(r.height for r in rows)
#             final_img = Image.new("RGB", (total_width, total_height))
#             y_offset = 0
#             for r in rows:
#                 final_img.paste(r, (0, y_offset))
#                 y_offset += r.height
#         else:
#             final_img = Image.new("RGB", (400, 200), color="gray")

#         self.root.update_idletasks()
#         canvas_w = max(1, self.canvas.winfo_width())
#         canvas_h = max(1, self.canvas.winfo_height())
#         scale = min(canvas_w / final_img.width, canvas_h / final_img.height, 1)
#         new_size = (max(1, int(final_img.width * scale)), max(1, int(final_img.height * scale)))
#         final_img = final_img.resize(new_size, Image.Resampling.LANCZOS)

#         self.tk_image = ImageTk.PhotoImage(final_img)
#         self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
#         self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

#         self.root.title(f"Viewing: {folder}")
#         self.status_entry.delete(0, tk.END)
#         self.status_entry.insert(0, entry["status"] or "")

#     def next_folder(self):
#         if self.current_index < len(self.folder_names) - 1:
#             self.current_index += 1
#             self.display_images()

#     def prev_folder(self):
#         if self.current_index > 0:
#             self.current_index -= 1
#             self.display_images()

#     def search_folder(self):
#         name = self.search_entry.get().strip()
#         if name in self.folder_names:
#             self.current_index = self.folder_names.index(name)
#             self.display_images()
#         else:
#             self.root.title(f"Folder '{name}' not found")

#     def save_status(self):
#         folder = self.folder_names[self.current_index]
#         new_status = self.status_entry.get().strip()
#         self.data[folder]["status"] = new_status

#         for row in self.ws.iter_rows(min_row=2):
#             if row[0].value == folder:
#                 row[1].value = new_status
#                 break

#         self.wb.save(EXCEL_FILE)
#         self.root.title(f"Saved: {folder} - {new_status}")


# if __name__ == "__main__":
#     root = tk.Tk()
#     root.geometry("1000x600")
#     viewer = ImageViewer(root)
#     root.mainloop()
# # /////////////////////////////////////////////////////


# # import os
# # import tkinter as tk
# # from PIL import Image, ImageTk, ImageDraw, ImageFont
# # from openpyxl import Workbook, load_workbook
# # from openpyxl.utils.exceptions import InvalidFileException
# # import webbrowser
# # import requests
# # from io import BytesIO

# # # === Configuration ===
# # EXCEL_FILE = "folder_status.xlsx"
# # BASE_URL = "http://yourdomain.com/Bucket"  # Replace with your actual base URL

# # # === Folder Names (Simulated or actual list from a server, here using local list) ===
# # BUCKET_PATH = r"C:\Users\pc\Downloads\Bucket\Bucket"  # Used to discover folder list
# # timestamp_folders = sorted([
# #     f for f in os.listdir(BUCKET_PATH)
# #     if os.path.isdir(os.path.join(BUCKET_PATH, f))
# # ])

# # # === Excel Setup ===
# # if os.path.exists(EXCEL_FILE):
# #     try:
# #         wb = load_workbook(EXCEL_FILE)
# #         ws = wb.active
# #     except InvalidFileException:
# #         wb = Workbook()
# #         ws = wb.active
# #         ws.title = "Folder Status"
# #         ws.append(["Folder Name", "Status"])
# # else:
# #     wb = Workbook()
# #     ws = wb.active
# #     ws.title = "Folder Status"
# #     ws.append(["Folder Name", "Status"])

# # status_dict = {
# #     row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
# # }


# # class ImageViewer:
# #     def __init__(self, root):
# #         self.root = root
# #         self.root.title("Timestamp Image Viewer")
# #         self.current_index = 0

# #         # === Search Frame ===
# #         self.search_frame = tk.Frame(self.root)
# #         self.search_frame.pack(fill=tk.X, padx=5, pady=5)

# #         self.search_entry = tk.Entry(self.search_frame)
# #         self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

# #         self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
# #         self.search_button.pack(side=tk.LEFT)

# #         # === Canvas ===
# #         self.canvas = tk.Canvas(root, bg='black')
# #         self.canvas.pack(fill=tk.BOTH, expand=True)

# #         # === Navigation Buttons ===
# #         self.button_frame = tk.Frame(self.root)
# #         self.button_frame.pack(fill=tk.X, pady=10)

# #         self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
# #         self.prev_button.pack(side=tk.LEFT, padx=10)

# #         self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
# #         self.next_button.pack(side=tk.RIGHT, padx=10)

# #         # === Status Entry ===
# #         self.status_label = tk.Label(self.root, text="Enter Status:")
# #         self.status_label.pack(pady=5)

# #         self.status_entry = tk.Entry(self.root)
# #         self.status_entry.pack(fill=tk.X, padx=5, pady=5)
# #         self.status_entry.bind("<Return>", lambda event: self.save_status())

# #         self.save_button = tk.Button(self.root, text="Save Status", command=self.save_status)
# #         self.save_button.pack(pady=5)

# #         # === URL Display ===
# #         self.url_label = tk.Label(self.root, text="", fg="blue", cursor="hand2")
# #         self.url_label.pack(pady=5)
# #         self.url_label.bind("<Button-1>", self.open_url)

# #         # === Keyboard Shortcuts ===
# #         self.root.bind('<KeyPress>', self.key_handler)
# #         self.root.bind("<Button-1>", self.on_root_click)

# #         # === Display Initial Images ===
# #         self.display_images()

# #     def key_handler(self, event):
# #         if self.root.focus_get() in (self.search_entry, self.status_entry):
# #             return
# #         if event.keysym.lower() == 'd':
# #             self.next_folder()
# #         elif event.keysym.lower() == 'a':
# #             self.prev_folder()

# #     def on_root_click(self, event):
# #         widget = event.widget
# #         if widget not in (self.search_entry, self.status_entry):
# #             self.search_entry.selection_clear()
# #             self.root.focus()

# #     def get_image_row_from_urls(self, folder_name, subfolder):
# #         base_folder_url = f"{BASE_URL}/{folder_name}/{subfolder}"
# #         image_urls = [f"{base_folder_url}/{i:03d}.jpg" for i in range(5)]

# #         images = []
# #         for url in image_urls:
# #             try:
# #                 response = requests.get(url)
# #                 response.raise_for_status()
# #                 img = Image.open(BytesIO(response.content)).resize((200, 200))
# #                 images.append(img)
# #             except Exception as e:
# #                 print(f"Failed to load {url}: {e}")
# #         return images

# #     def label_row(self, text, width):
# #         img = Image.new("RGB", (width, 30), color="gray")
# #         draw = ImageDraw.Draw(img)
# #         font = ImageFont.load_default()
# #         bbox = draw.textbbox((0, 0), text, font=font)
# #         w = bbox[2] - bbox[0]
# #         h = bbox[3] - bbox[1]
# #         draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
# #         return img

# #     def get_url_from_path(self, folder_name):
# #         return f"{BASE_URL}/{folder_name}"

# #     def open_url(self, event):
# #         url = self.url_label.cget("text")
# #         if url:
# #             webbrowser.open_new(url)

# #     def display_images(self):
# #         self.canvas.delete("all")
# #         folder_name = timestamp_folders[self.current_index]

# #         input_images = self.get_image_row_from_urls(folder_name, "input")
# #         predicted_images = self.get_image_row_from_urls(folder_name, "predicted")

# #         rows = []
# #         for label, images in zip(["Input Images", "Predicted Images"], [input_images, predicted_images]):
# #             if images:
# #                 row_width = sum(img.width for img in images)
# #                 row_height = max(img.height for img in images)
# #                 row_image = Image.new("RGB", (row_width, row_height))
# #                 x_offset = 0
# #                 for img in images:
# #                     row_image.paste(img, (x_offset, 0))
# #                     x_offset += img.width

# #                 label_image = self.label_row(label, row_width)
# #                 combined = Image.new("RGB", (row_width, row_height + 30), color="black")
# #                 combined.paste(label_image, (0, 0))
# #                 combined.paste(row_image, (0, 30))
# #                 rows.append(combined)

# #         if rows:
# #             total_width = max(r.width for r in rows)
# #             total_height = sum(r.height for r in rows)
# #             final_image = Image.new("RGB", (total_width, total_height))
# #             y_offset = 0
# #             for r in rows:
# #                 final_image.paste(r, (0, y_offset))
# #                 y_offset += r.height
# #         else:
# #             final_image = Image.new("RGB", (400, 200), color='gray')

# #         self.root.update_idletasks()
# #         canvas_width = max(1, self.canvas.winfo_width())
# #         canvas_height = max(1, self.canvas.winfo_height())
# #         if final_image.width > 0 and final_image.height > 0:
# #             scale_w = canvas_width / final_image.width
# #             scale_h = canvas_height / final_image.height
# #             scale = min(scale_w, scale_h, 1.0)
# #             new_size = (max(1, int(final_image.width * scale)), max(1, int(final_image.height * scale)))
# #             final_image = final_image.resize(new_size, Image.Resampling.LANCZOS)

# #         self.tk_image = ImageTk.PhotoImage(final_image)
# #         self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
# #         self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

# #         self.root.title(f"Viewing: {folder_name}")
# #         self.status_entry.delete(0, tk.END)
# #         self.status_entry.insert(0, status_dict.get(folder_name, ""))

# #         self.url_label.config(text=self.get_url_from_path(folder_name))

# #     def next_folder(self, event=None):
# #         if self.current_index < len(timestamp_folders) - 1:
# #             self.current_index += 1
# #             self.display_images()

# #     def prev_folder(self, event=None):
# #         if self.current_index > 0:
# #             self.current_index -= 1
# #             self.display_images()

# #     def search_folder(self):
# #         folder_name = self.search_entry.get().strip()
# #         if folder_name in timestamp_folders:
# #             self.current_index = timestamp_folders.index(folder_name)
# #             self.display_images()
# #         else:
# #             self.root.title(f"Folder '{folder_name}' not found")

# #     def save_status(self):
# #         folder_name = timestamp_folders[self.current_index]
# #         status = self.status_entry.get().strip()
# #         if status:
# #             status_dict[folder_name] = status
# #             updated = False
# #             for row in ws.iter_rows(min_row=2):
# #                 if row[0].value == folder_name:
# #                     row[1].value = status
# #                     updated = True
# #                     break
# #             if not updated:
# #                 ws.append([folder_name, status])
# #             wb.save(EXCEL_FILE)
# #             self.root.title(f"Saved: {folder_name} - {status}")
# #         else:
# #             self.root.title("Enter a status before saving")


# # if __name__ == "__main__":
# #     root = tk.Tk()
# #     root.geometry("1000x600")
# #     viewer = ImageViewer(root)
# #     root.mainloop()



import tkinter as tk
from PIL import Image, ImageTk, ImageDraw, ImageFont
from openpyxl import load_workbook
import requests
from io import BytesIO

EXCEL_FILE = "folder_status.xlsx"

# === Load Data from Excel ===
def load_excel_data():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        folder_name = row[0]
        status = row[1]
        input_urls = row[2].split(",") if row[2] else []
        predicted_urls = row[3].split(",") if row[3] else []
        data[folder_name] = {
            "status": status,
            "input_urls": input_urls,
            "predicted_urls": predicted_urls
        }
    return data, wb, ws

class ImageViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("URL Image Viewer")
        self.current_index = 0

        self.data, self.wb, self.ws = load_excel_data()
        self.folder_names = list(self.data.keys())

        # === GUI Layout ===
        self.search_frame = tk.Frame(root)
        self.search_frame.pack(fill=tk.X, padx=5, pady=5)

        self.search_entry = tk.Entry(self.search_frame)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
        self.search_button.pack(side=tk.LEFT, padx=5)

        self.canvas = tk.Canvas(root, bg='black')
        self.canvas.pack(fill=tk.BOTH, expand=True)

        self.button_frame = tk.Frame(root)
        self.button_frame.pack(fill=tk.X, pady=10)

        self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
        self.prev_button.pack(side=tk.LEFT, padx=10)

        self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
        self.next_button.pack(side=tk.RIGHT, padx=10)

        self.status_label = tk.Label(root, text="Enter Status:")
        self.status_label.pack(pady=5)

        self.status_entry = tk.Entry(root)
        self.status_entry.pack(fill=tk.X, padx=5, pady=5)
        self.status_entry.bind("<Return>", lambda e: self.save_status())

        self.save_button = tk.Button(root, text="Save Status", command=self.save_status)
        self.save_button.pack(pady=5)

        self.display_images()

    def load_images_from_urls(self, urls):
        images = []
        for url in urls:
            try:
                resp = requests.get(url)
                resp.raise_for_status()
                img = Image.open(BytesIO(resp.content)).resize((200, 200))
                images.append(img)
            except Exception as e:
                print(f"Error loading image from {url}: {e}")
        return images

    def label_row(self, text, width):
        img = Image.new("RGB", (width, 30), color="gray")
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
        return img

    def display_images(self):
        self.canvas.delete("all")
        folder = self.folder_names[self.current_index]
        entry = self.data[folder]

        input_imgs = self.load_images_from_urls(entry["input_urls"])
        predicted_imgs = self.load_images_from_urls(entry["predicted_urls"])

        rows = []
        for label, images in zip(["Input Images", "Predicted Images"], [input_imgs, predicted_imgs]):
            if images:
                row_width = sum(img.width for img in images)
                row_height = max(img.height for img in images)
                row_img = Image.new("RGB", (row_width, row_height))
                x_offset = 0
                for img in images:
                    row_img.paste(img, (x_offset, 0))
                    x_offset += img.width

                label_img = self.label_row(label, row_width)
                combined = Image.new("RGB", (row_width, row_height + 30), color="black")
                combined.paste(label_img, (0, 0))
                combined.paste(row_img, (0, 30))
                rows.append(combined)

        if rows:
            total_width = max(r.width for r in rows)
            total_height = sum(r.height for r in rows)
            final_img = Image.new("RGB", (total_width, total_height))
            y_offset = 0
            for r in rows:
                final_img.paste(r, (0, y_offset))
                y_offset += r.height
        else:
            final_img = Image.new("RGB", (400, 200), color="gray")

        self.root.update_idletasks()
        canvas_w = max(1, self.canvas.winfo_width())
        canvas_h = max(1, self.canvas.winfo_height())
        scale = min(canvas_w / final_img.width, canvas_h / final_img.height, 1)
        new_size = (max(1, int(final_img.width * scale)), max(1, int(final_img.height * scale)))
        final_img = final_img.resize(new_size, Image.Resampling.LANCZOS)

        self.tk_image = ImageTk.PhotoImage(final_img)
        self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
        self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

        self.root.title(f"Viewing: {folder}")
        self.status_entry.delete(0, tk.END)
        self.status_entry.insert(0, entry["status"] or "")

    def next_folder(self):
        if self.current_index < len(self.folder_names) - 1:
            self.current_index += 1
            self.display_images()

    def prev_folder(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.display_images()

    def search_folder(self):
        name = self.search_entry.get().strip()
        if name in self.folder_names:
            self.current_index = self.folder_names.index(name)
            self.display_images()
        else:
            self.root.title(f"Folder '{name}' not found")

    def save_status(self):
        folder = self.folder_names[self.current_index]
        new_status = self.status_entry.get().strip()
        self.data[folder]["status"] = new_status

        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == folder:
                row[1].value = new_status
                break

        self.wb.save(EXCEL_FILE)
        self.root.title(f"Saved: {folder} - {new_status}")


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1000x600")
    viewer = ImageViewer(root)
    root.mainloop()
