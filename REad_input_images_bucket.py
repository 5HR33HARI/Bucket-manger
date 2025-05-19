# import os
# import tkinter as tk
# from PIL import Image, ImageTk, ImageDraw, ImageFont
# from openpyxl import Workbook, load_workbook
# from openpyxl.utils.exceptions import InvalidFileException

# # === Paths ===
# EXCEL_FILE = "folder_status.xlsx"
# BUCKET_PATH = r"C:\Users\pc\Downloads\Bucket\Bucket"

# # === Folder List ===
# timestamp_folders = sorted([
#     f for f in os.listdir(BUCKET_PATH)
#     if os.path.isdir(os.path.join(BUCKET_PATH, f))
# ])

# # === Excel Setup ===
# if os.path.exists(EXCEL_FILE):
#     try:
#         wb = load_workbook(EXCEL_FILE)
#         ws = wb.active
#     except InvalidFileException:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Folder Status"
#         ws.append(["Folder Name", "Status"])
# else:
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Folder Status"
#     ws.append(["Folder Name", "Status"])

# # === Dictionary for Existing Statuses ===
# status_dict = {
#     row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
# }


# class ImageViewer:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Timestamp Image Viewer")
#         self.current_index = 0

#         # === Search Frame ===
#         self.search_frame = tk.Frame(self.root)
#         self.search_frame.pack(fill=tk.X, padx=5, pady=5)

#         self.search_entry = tk.Entry(self.search_frame)
#         self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

#         self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
#         self.search_button.pack(side=tk.LEFT)

#         # === Canvas ===
#         self.canvas = tk.Canvas(root, bg='black')
#         self.canvas.pack(fill=tk.BOTH, expand=True)

#         # === Navigation Buttons ===
#         self.button_frame = tk.Frame(self.root)
#         self.button_frame.pack(fill=tk.X, pady=10)

#         self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
#         self.prev_button.pack(side=tk.LEFT, padx=10)

#         self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
#         self.next_button.pack(side=tk.RIGHT, padx=10)

#         # === Status Entry ===
#         self.status_label = tk.Label(self.root, text="Enter Status:")
#         self.status_label.pack(pady=5)

#         self.status_entry = tk.Entry(self.root)
#         self.status_entry.pack(fill=tk.X, padx=5, pady=5)

#         self.save_button = tk.Button(self.root, text="Save Status", command=self.save_status)
#         self.save_button.pack(pady=5)

#         # === Keyboard Shortcuts ===
#         self.root.bind('<KeyPress>', self.key_handler)

#         # === Mouse click to disable search focus ===
#         self.root.bind("<Button-1>", self.on_root_click)

#         # === Display Initial ===
#         self.display_images()

#     def key_handler(self, event):
#         if self.root.focus_get() in (self.search_entry, self.status_entry):
#             return
#         if event.keysym.lower() == 'd':
#             self.next_folder()
#         elif event.keysym.lower() == 'a':
#             self.prev_folder()

#     def on_root_click(self, event):
#         widget = event.widget
#         if widget != self.search_entry:
#             self.search_entry.selection_clear()
#             self.root.focus()

#     def get_image_row(self, folder_path):
#         if not os.path.exists(folder_path):
#             return []
#         return sorted([
#             os.path.join(folder_path, f)
#             for f in os.listdir(folder_path)
#             if f.endswith('.jpg')
#         ])[:5]

#     def label_row(self, text, width):
#         img = Image.new("RGB", (width, 30), color="gray")
#         draw = ImageDraw.Draw(img)
#         font = ImageFont.load_default()
#         bbox = draw.textbbox((0, 0), text, font=font)
#         w = bbox[2] - bbox[0]
#         h = bbox[3] - bbox[1]
#         draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
#         return img

#     def display_images(self):
#         self.canvas.delete("all")
#         folder_name = timestamp_folders[self.current_index]
#         folder_path = os.path.join(BUCKET_PATH, folder_name)

#         input_images = self.get_image_row(os.path.join(folder_path, "input"))
#         predicted_images = self.get_image_row(os.path.join(folder_path, "predicted"))

#         pil_inputs = [Image.open(path).resize((200, 200)) for path in input_images]
#         pil_predicts = [Image.open(path).resize((200, 200)) for path in predicted_images]

#         rows = []
#         for label, images in zip(["Input Images", "Predicted Images"], [pil_inputs, pil_predicts]):
#             if images:
#                 row_width = sum(img.width for img in images)
#                 row_height = max(img.height for img in images)
#                 row_image = Image.new("RGB", (row_width, row_height))
#                 x_offset = 0
#                 for img in images:
#                     row_image.paste(img, (x_offset, 0))
#                     x_offset += img.width

#                 label_image = self.label_row(label, row_width)
#                 combined = Image.new("RGB", (row_width, row_height + 30), color="black")
#                 combined.paste(label_image, (0, 0))
#                 combined.paste(row_image, (0, 30))
#                 rows.append(combined)

#         if rows:
#             total_width = max(r.width for r in rows)
#             total_height = sum(r.height for r in rows)
#             final_image = Image.new("RGB", (total_width, total_height))
#             y_offset = 0
#             for r in rows:
#                 final_image.paste(r, (0, y_offset))
#                 y_offset += r.height
#         else:
#             final_image = Image.new("RGB", (400, 200), color='gray')

#         # Resize image to fit canvas
#         self.root.update_idletasks()
#         canvas_width = max(1, self.canvas.winfo_width())
#         canvas_height = max(1, self.canvas.winfo_height())
#         if final_image.width > 0 and final_image.height > 0:
#             scale_w = canvas_width / final_image.width
#             scale_h = canvas_height / final_image.height
#             scale = min(scale_w, scale_h, 1.0)
#             new_size = (max(1, int(final_image.width * scale)), max(1, int(final_image.height * scale)))
#             final_image = final_image.resize(new_size, Image.Resampling.LANCZOS)

#         self.tk_image = ImageTk.PhotoImage(final_image)
#         self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
#         self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

#         self.root.title(f"Viewing: {folder_name}")

#         # Auto-fill status
#         self.status_entry.delete(0, tk.END)
#         self.status_entry.insert(0, status_dict.get(folder_name, ""))

#     def next_folder(self, event=None):
#         if self.current_index < len(timestamp_folders) - 1:
#             self.current_index += 1
#             self.display_images()

#     def prev_folder(self, event=None):
#         if self.current_index > 0:
#             self.current_index -= 1
#             self.display_images()

#     def search_folder(self):
#         folder_name = self.search_entry.get().strip()
#         if folder_name in timestamp_folders:
#             self.current_index = timestamp_folders.index(folder_name)
#             self.display_images()
#         else:
#             self.root.title(f"Folder '{folder_name}' not found")

#     def save_status(self):
#         folder_name = timestamp_folders[self.current_index]
#         status = self.status_entry.get().strip()
#         if status:
#             status_dict[folder_name] = status

#             # Update or append in Excel
#             updated = False
#             for row in ws.iter_rows(min_row=2):
#                 if row[0].value == folder_name:
#                     row[1].value = status
#                     updated = True
#                     break
#             if not updated:
#                 ws.append([folder_name, status])

#             wb.save(EXCEL_FILE)
#             self.root.title(f"Saved: {folder_name} - {status}")
#         else:
#             self.root.title("Enter a status before saving")


# if __name__ == "__main__":
#     root = tk.Tk()
#     root.geometry("1000x600")
#     viewer = ImageViewer(root)
#     root.mainloop()


#######################################################################################

# import os 
# import tkinter as tk
# from PIL import Image, ImageTk, ImageDraw, ImageFont
# from openpyxl import Workbook, load_workbook
# from openpyxl.utils.exceptions import InvalidFileException

# # === Paths ===
# EXCEL_FILE = "folder_status.xlsx"
# BUCKET_PATH = r"C:\Users\pc\Downloads\Bucket\Bucket"

# # === Folder List ===
# timestamp_folders = sorted([
#     f for f in os.listdir(BUCKET_PATH)
#     if os.path.isdir(os.path.join(BUCKET_PATH, f))
# ])

# # === Excel Setup ===
# if os.path.exists(EXCEL_FILE):
#     try:
#         wb = load_workbook(EXCEL_FILE)
#         ws = wb.active
#     except InvalidFileException:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Folder Status"
#         ws.append(["Folder Name", "Status"])
# else:
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Folder Status"
#     ws.append(["Folder Name", "Status"])

# # === Dictionary for Existing Statuses ===
# status_dict = {
#     row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
# }


# class ImageViewer:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Timestamp Image Viewer")
#         self.current_index = 0

#         # === Search Frame ===
#         self.search_frame = tk.Frame(self.root)
#         self.search_frame.pack(fill=tk.X, padx=5, pady=5)

#         self.search_entry = tk.Entry(self.search_frame)
#         self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

#         self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
#         self.search_button.pack(side=tk.LEFT)

#         # === Canvas ===
#         self.canvas = tk.Canvas(root, bg='black')
#         self.canvas.pack(fill=tk.BOTH, expand=True)

#         # === Navigation Buttons ===
#         self.button_frame = tk.Frame(self.root)
#         self.button_frame.pack(fill=tk.X, pady=10)

#         self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
#         self.prev_button.pack(side=tk.LEFT, padx=10)

#         self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
#         self.next_button.pack(side=tk.RIGHT, padx=10)

#         # === Status Entry ===
#         self.status_label = tk.Label(self.root, text="Enter Status:")
#         self.status_label.pack(pady=5)

#         self.status_entry = tk.Entry(self.root)
#         self.status_entry.pack(fill=tk.X, padx=5, pady=5)

#         self.save_button = tk.Button(self.root, text="Save Status", command=self.save_status)
#         self.save_button.pack(pady=5)

#         # === Keyboard Shortcuts ===
#         self.root.bind('<KeyPress>', self.key_handler)

#         # === Mouse click to disable search focus only ===
#         self.root.bind("<Button-1>", self.on_root_click)

#         # === Display Initial ===
#         self.display_images()

#     def key_handler(self, event):
#         if self.root.focus_get() in (self.search_entry, self.status_entry):
#             return
#         if event.keysym.lower() == 'd':
#             self.next_folder()
#         elif event.keysym.lower() == 'a':
#             self.prev_folder()

#     def on_root_click(self, event):
#         widget = event.widget
#         # Only unfocus search_entry
#         if widget != self.search_entry:
#             self.search_entry.selection_clear()
#             self.root.focus()

#     def get_image_row(self, folder_path):
#         if not os.path.exists(folder_path):
#             return []
#         return sorted([
#             os.path.join(folder_path, f)
#             for f in os.listdir(folder_path)
#             if f.endswith('.jpg')
#         ])[:5]

#     def label_row(self, text, width):
#         img = Image.new("RGB", (width, 30), color="gray")
#         draw = ImageDraw.Draw(img)
#         font = ImageFont.load_default()
#         bbox = draw.textbbox((0, 0), text, font=font)
#         w = bbox[2] - bbox[0]
#         h = bbox[3] - bbox[1]
#         draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
#         return img

#     def display_images(self):
#         self.canvas.delete("all")
#         folder_name = timestamp_folders[self.current_index]
#         folder_path = os.path.join(BUCKET_PATH, folder_name)

#         input_images = self.get_image_row(os.path.join(folder_path, "input"))
#         predicted_images = self.get_image_row(os.path.join(folder_path, "predicted"))

#         pil_inputs = [Image.open(path).resize((200, 200)) for path in input_images]
#         pil_predicts = [Image.open(path).resize((200, 200)) for path in predicted_images]

#         rows = []
#         for label, images in zip(["Input Images", "Predicted Images"], [pil_inputs, pil_predicts]):
#             if images:
#                 row_width = sum(img.width for img in images)
#                 row_height = max(img.height for img in images)
#                 row_image = Image.new("RGB", (row_width, row_height))
#                 x_offset = 0
#                 for img in images:
#                     row_image.paste(img, (x_offset, 0))
#                     x_offset += img.width

#                 label_image = self.label_row(label, row_width)
#                 combined = Image.new("RGB", (row_width, row_height + 30), color="black")
#                 combined.paste(label_image, (0, 0))
#                 combined.paste(row_image, (0, 30))
#                 rows.append(combined)

#         if rows:
#             total_width = max(r.width for r in rows)
#             total_height = sum(r.height for r in rows)
#             final_image = Image.new("RGB", (total_width, total_height))
#             y_offset = 0
#             for r in rows:
#                 final_image.paste(r, (0, y_offset))
#                 y_offset += r.height
#         else:
#             final_image = Image.new("RGB", (400, 200), color='gray')

#         # Resize image to fit canvas
#         self.root.update_idletasks()
#         canvas_width = max(1, self.canvas.winfo_width())
#         canvas_height = max(1, self.canvas.winfo_height())
#         if final_image.width > 0 and final_image.height > 0:
#             scale_w = canvas_width / final_image.width
#             scale_h = canvas_height / final_image.height
#             scale = min(scale_w, scale_h, 1.0)
#             new_size = (max(1, int(final_image.width * scale)), max(1, int(final_image.height * scale)))
#             final_image = final_image.resize(new_size, Image.Resampling.LANCZOS)

#         self.tk_image = ImageTk.PhotoImage(final_image)
#         self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
#         self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

#         self.root.title(f"Viewing: {folder_name}")

#         # Auto-fill status
#         self.status_entry.delete(0, tk.END)
#         self.status_entry.insert(0, status_dict.get(folder_name, ""))

#     def next_folder(self, event=None):
#         if self.current_index < len(timestamp_folders) - 1:
#             self.current_index += 1
#             self.display_images()

#     def prev_folder(self, event=None):
#         if self.current_index > 0:
#             self.current_index -= 1
#             self.display_images()

#     def search_folder(self):
#         folder_name = self.search_entry.get().strip()
#         if folder_name in timestamp_folders:
#             self.current_index = timestamp_folders.index(folder_name)
#             self.display_images()
#         else:
#             self.root.title(f"Folder '{folder_name}' not found")

#     def save_status(self):
#         folder_name = timestamp_folders[self.current_index]
#         status = self.status_entry.get().strip()
#         if status:
#             status_dict[folder_name] = status

#             # Update or append in Excel
#             updated = False
#             for row in ws.iter_rows(min_row=2):
#                 if row[0].value == folder_name:
#                     row[1].value = status
#                     updated = True
#                     break
#             if not updated:
#                 ws.append([folder_name, status])

#             wb.save(EXCEL_FILE)
#             self.root.title(f"Saved: {folder_name} - {status}")
#         else:
#             self.root.title("Enter a status before saving")


# if __name__ == "__main__":
#     root = tk.Tk()
#     root.geometry("1000x600")
#     viewer = ImageViewer(root)
#     root.mainloop()

# /////////////////////////////////////////////////////////////////////////////

import os 
import tkinter as tk
from PIL import Image, ImageTk, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# === Paths ===
EXCEL_FILE = "folder_status.xlsx"
BUCKET_PATH = r"C:\Users\pc\Downloads\Bucket\Bucket"

# === Folder List ===
timestamp_folders = sorted([
    f for f in os.listdir(BUCKET_PATH)
    if os.path.isdir(os.path.join(BUCKET_PATH, f))
])

# === Excel Setup ===
if os.path.exists(EXCEL_FILE):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except InvalidFileException:
        wb = Workbook()
        ws = wb.active
        ws.title = "Folder Status"
        ws.append(["Folder Name", "Status"])
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Status"
    ws.append(["Folder Name", "Status"])

# === Dictionary for Existing Statuses ===
status_dict = {
    row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
}

class ImageViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Timestamp Image Viewer")
        self.current_index = 0

        # === Search Frame ===
        self.search_frame = tk.Frame(self.root)
        self.search_frame.pack(fill=tk.X, padx=5, pady=5)

        self.search_entry = tk.Entry(self.search_frame)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        self.search_button = tk.Button(self.search_frame, text="Search", command=self.search_folder)
        self.search_button.pack(side=tk.LEFT)

        # === Canvas ===
        self.canvas = tk.Canvas(root, bg='black')
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # === Navigation Buttons ===
        self.button_frame = tk.Frame(self.root)
        self.button_frame.pack(fill=tk.X, pady=10)

        self.prev_button = tk.Button(self.button_frame, text="⟵ Previous", command=self.prev_folder)
        self.prev_button.pack(side=tk.LEFT, padx=10)

        self.next_button = tk.Button(self.button_frame, text="Next ⟶", command=self.next_folder)
        self.next_button.pack(side=tk.RIGHT, padx=10)

        # === Status Entry ===
        self.status_label = tk.Label(self.root, text="Enter Status:")
        self.status_label.pack(pady=5)

        self.status_entry = tk.Entry(self.root)
        self.status_entry.pack(fill=tk.X, padx=5, pady=5)

        # ⏎ Save on Enter in status entry
        self.status_entry.bind("<Return>", lambda event: self.save_status())

        self.save_button = tk.Button(self.root, text="Save Status", command=self.save_status)
        self.save_button.pack(pady=5)

        # === Keyboard Shortcuts ===
        self.root.bind('<KeyPress>', self.key_handler)

        # === Mouse click to disable search bar editing mode ===
        self.root.bind("<Button-1>", self.on_root_click)

        # === Display Initial ===
        self.display_images()

    def key_handler(self, event):
        if self.root.focus_get() in (self.search_entry, self.status_entry):
            return
        if event.keysym.lower() == 'd':
            self.next_folder()
        elif event.keysym.lower() == 'a':
            self.prev_folder()

    def on_root_click(self, event):
        widget = event.widget
        # Only unfocus search entry if the user clicks outside both search and status entries
        if widget not in (self.search_entry, self.status_entry):
            self.search_entry.selection_clear()
            self.root.focus()

    def get_image_row(self, folder_path):
        if not os.path.exists(folder_path):
            return []
        return sorted([
            os.path.join(folder_path, f)
            for f in os.listdir(folder_path)
            if f.endswith('.jpg')
        ])[:5]

    def label_row(self, text, width):
        img = Image.new("RGB", (width, 30), color="gray")
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        draw.text(((width - w) / 2, (30 - h) / 2), text, fill="white", font=font)
        return img

    def display_images(self):
        self.canvas.delete("all")
        folder_name = timestamp_folders[self.current_index]
        folder_path = os.path.join(BUCKET_PATH, folder_name)

        input_images = self.get_image_row(os.path.join(folder_path, "input"))
        predicted_images = self.get_image_row(os.path.join(folder_path, "predicted"))

        pil_inputs = [Image.open(path).resize((200, 200)) for path in input_images]
        pil_predicts = [Image.open(path).resize((200, 200)) for path in predicted_images]

        rows = []
        for label, images in zip(["Input Images", "Predicted Images"], [pil_inputs, pil_predicts]):
            if images:
                row_width = sum(img.width for img in images)
                row_height = max(img.height for img in images)
                row_image = Image.new("RGB", (row_width, row_height))
                x_offset = 0
                for img in images:
                    row_image.paste(img, (x_offset, 0))
                    x_offset += img.width

                label_image = self.label_row(label, row_width)
                combined = Image.new("RGB", (row_width, row_height + 30), color="black")
                combined.paste(label_image, (0, 0))
                combined.paste(row_image, (0, 30))
                rows.append(combined)

        if rows:
            total_width = max(r.width for r in rows)
            total_height = sum(r.height for r in rows)
            final_image = Image.new("RGB", (total_width, total_height))
            y_offset = 0
            for r in rows:
                final_image.paste(r, (0, y_offset))
                y_offset += r.height
        else:
            final_image = Image.new("RGB", (400, 200), color='gray')

        # Resize image to fit canvas
        self.root.update_idletasks()
        canvas_width = max(1, self.canvas.winfo_width())
        canvas_height = max(1, self.canvas.winfo_height())
        if final_image.width > 0 and final_image.height > 0:
            scale_w = canvas_width / final_image.width
            scale_h = canvas_height / final_image.height
            scale = min(scale_w, scale_h, 1.0)
            new_size = (max(1, int(final_image.width * scale)), max(1, int(final_image.height * scale)))
            final_image = final_image.resize(new_size, Image.Resampling.LANCZOS)

        self.tk_image = ImageTk.PhotoImage(final_image)
        self.canvas.create_image(0, 0, anchor='nw', image=self.tk_image)
        self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

        self.root.title(f"Viewing: {folder_name}")

        # Auto-fill status
        self.status_entry.delete(0, tk.END)
        self.status_entry.insert(0, status_dict.get(folder_name, ""))

    def next_folder(self, event=None):
        if self.current_index < len(timestamp_folders) - 1:
            self.current_index += 1
            self.display_images()

    def prev_folder(self, event=None):
        if self.current_index > 0:
            self.current_index -= 1
            self.display_images()

    def search_folder(self):
        folder_name = self.search_entry.get().strip()
        if folder_name in timestamp_folders:
            self.current_index = timestamp_folders.index(folder_name)
            self.display_images()
        else:
            self.root.title(f"Folder '{folder_name}' not found")

    def save_status(self):
        folder_name = timestamp_folders[self.current_index]
        status = self.status_entry.get().strip()
        if status:
            status_dict[folder_name] = status

            # Update or append in Excel
            updated = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == folder_name:
                    row[1].value = status
                    updated = True
                    break
            if not updated:
                ws.append([folder_name, status])

            wb.save(EXCEL_FILE)
            self.root.title(f"Saved: {folder_name} - {status}")
        else:
            self.root.title("Enter a status before saving")


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1000x600")
    viewer = ImageViewer(root)
    root.mainloop()
